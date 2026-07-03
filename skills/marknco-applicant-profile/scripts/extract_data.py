#!/usr/bin/env python3
"""
extract_data.py — 마크앤컴퍼니 프로필 스킬용 엑셀 데이터 추출기
Usage:
  python3 extract_data.py \
    --xlsx 파일.xlsx \
    --sheet "정리 시트" \
    --name-col 4 \
    --data-cols "4-22" \
    --output companies.json \
    [--ext-sheet "innoclaw 데이터"] \
    [--ext-name-col 3] \
    [--ext-data-cols "3-17"] \
    [--ext-output ext_data.json] \
    [--skip-keywords "취소,불성실"]
"""

import argparse
import json
import sys
import openpyxl


def parse_col_range(s):
    """'4-22' → list(range(4,23)), '3,5,7' → [3,5,7], '4' → [4]"""
    s = s.strip()
    if "-" in s and "," not in s:
        a, b = s.split("-")
        return list(range(int(a), int(b) + 1))
    elif "," in s:
        return [int(x) for x in s.split(",")]
    else:
        return [int(s)]


def extract_sheet(ws, name_col, data_cols, skip_keywords=None):
    skip_keywords = skip_keywords or []
    headers = [ws.cell(row=1, column=c).value for c in data_cols]
    records = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=name_col).value
        if not name:
            continue
        # 스크리닝 컬럼 건너뛰기 (A, B 컬럼 체크)
        if skip_keywords:
            skip = False
            for col in range(1, min(4, ws.max_column + 1)):
                cell_val = str(ws.cell(row=row, column=col).value or "")
                if any(kw in cell_val for kw in skip_keywords):
                    skip = True
                    break
            if skip:
                continue
        rec = {}
        for i, col in enumerate(data_cols):
            h = headers[i]
            if h is None:
                continue
            v = ws.cell(row=row, column=col).value
            rec[h] = v
        records.append(rec)
    return records


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", required=True)
    p.add_argument("--sheet", required=True)
    p.add_argument("--name-col", type=int, required=True)
    p.add_argument("--data-cols", required=True)
    p.add_argument("--output", required=True)
    p.add_argument("--ext-sheet")
    p.add_argument("--ext-name-col", type=int)
    p.add_argument("--ext-data-cols")
    p.add_argument("--ext-output")
    p.add_argument("--skip-keywords", default="취소,불성실")
    args = p.parse_args()

    print(f"📂 엑셀 읽는 중: {args.xlsx}")
    wb = openpyxl.load_workbook(args.xlsx, data_only=True)

    if args.sheet not in wb.sheetnames:
        print(f"❌ 시트를 찾을 수 없음: '{args.sheet}'")
        print(f"   사용 가능한 시트: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[args.sheet]
    data_cols = parse_col_range(args.data_cols)
    skip_kw = [kw.strip() for kw in args.skip_keywords.split(",") if kw.strip()] if args.skip_keywords else []

    records = extract_sheet(ws, args.name_col, data_cols, skip_kw)
    print(f"✅ 메인 시트 '{args.sheet}': {len(records)}개 기업 추출")
    for r in records:
        name_field = list(r.keys())[0]
        print(f"   - {r.get(name_field, '?')}")

    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2, default=str)
    print(f"💾 저장: {args.output}")

    # 외부 데이터 시트
    if args.ext_sheet:
        if args.ext_sheet not in wb.sheetnames:
            print(f"⚠️  외부 데이터 시트 '{args.ext_sheet}' 없음 — 건너뜀")
            return
        ws2 = wb[args.ext_sheet]
        ext_cols = parse_col_range(args.ext_data_cols)
        ext_records = extract_sheet(ws2, args.ext_name_col, ext_cols, skip_keywords=[])

        # 기업명 기준 dict로 변환
        ext_headers = [ws2.cell(row=1, column=c).value for c in ext_cols]
        name_header = ext_headers[0]  # ext_name_col 기준 첫 번째 헤더
        ext_dict = {}
        for r in ext_records:
            key = r.get(name_header)
            if key:
                ext_dict[str(key)] = r

        # 의미있는 데이터 있는 기업만 표시
        meaningful = [k for k, v in ext_dict.items()
                      if any(val not in (None, "-", "") for val in list(v.values())[1:])]
        print(f"✅ 외부 시트 '{args.ext_sheet}': {len(ext_dict)}개 기업 중 데이터 있음: {meaningful}")

        ext_out = args.ext_output or args.output.replace(".json", "_ext.json")
        with open(ext_out, "w", encoding="utf-8") as f:
            json.dump(ext_dict, f, ensure_ascii=False, indent=2, default=str)
        print(f"💾 저장: {ext_out}")


if __name__ == "__main__":
    main()
